# This script is written to fix the naming convention of audio files.
# There are 2 main types of naming that has to be fixed, first is <speaker_name><speaker_id><file_id>.WAV. Second is <speaker_id><file_id>.WAV
# These file names should be fixed to <speaker_id><file_id>.wav
# This script also contains function to moves all files in a directory and it's subdirectory into another directory

import os
import openpyxl
import shutil
import random
import sys
folderName = "C:/Users/jazzt/Desktop/CCA/Deepfake_Audio_Detection/Mangio-RVC-v23.7.0_INFER_TRAIN/Mangio-RVC/opt/New_folder"

def list_folders_and_files(dir_path, funcPtr): # the function pointer should only take in the file path
    # Loop through all files and directories in the given path
    for item in os.listdir(dir_path):
        # Get the full path of the item
        item_path = dir_path + "/" + item
        
        if os.path.isdir(item_path):
            # If the item is a directory, recursively list its contents
            print(f"Entering folder: {item_path}")
            list_folders_and_files(item_path, funcPtr)
        else:
            # If the item is a file, print its path
            funcPtr(item_path)

def FixAudioFileName(filepath):
    if filepath[-3:].lower() != "wav":
        print(f"{filepath} is not a wav file")
        return
    dir, name = os.path.split(filepath)
    
    # fix first error, removing SPEAKER<number>
    if (not name[0].isdigit()):
        name = name[11:]
        
    
    # fix second error, where there are more than one .wav in the name, or wav is WAV
    wavOccurance = name.lower().count(".wav")
    while (wavOccurance != 0):
        name = name[:-4]
        wavOccurance-=1
    name = name + ".wav"

    # # fix third error, where there are more than 8 digits in the number
    # if len(name) > 12:
    #     name = name[1:]
    
    print(dir+"/"+name)
    os.rename(filepath, dir+"/"+name)

list_folders_and_files(folderName, FixAudioFileName)

# SET THE VC MODEL ID HERE
VC_id = 0

# Converts a spoof data with bonafide naming to spoof naming
def SetSpoofedDataName(startingPath): 
    pitchTable = ExcelSheetToTable(startingPath + "/voice_adjustment.xlsx")
    print(pitchTable)
    print(startingPath + "/voice_adjustment.xlsx")
    for subfolder in os.listdir(startingPath):
        if not os.path.isdir(startingPath+'/'+subfolder):
            print(subfolder)
            continue
        # Get the full path of the item
        target_id = subfolder[-4:]
        for fromFolder in os.listdir(startingPath+'/'+subfolder):
            source_id = fromFolder[-4:]
            for name in os.listdir(startingPath+'/'+subfolder+'/'+fromFolder):
                newname = target_id + name
                newname = newname[:8] + pitchTable[target_id][source_id] + f"{VC_id:02}" + newname[8:]
                # print(newname)
                # print(startingPath+'/'+subfolder+'/'+fromFolder+'/'+name, " -> ",
                #           startingPath+'/'+subfolder+'/'+fromFolder+'/'+newname)
                os.rename(startingPath+'/'+subfolder+'/'+fromFolder+'/'+name,
                          startingPath+'/'+subfolder+'/'+fromFolder+'/'+newname)

def ExcelSheetToTable(excelSheetPath): # converting excelsheet to a 2d array, 
    sheet = openpyxl.load_workbook(excelSheetPath)['Sheet1']
    
    table = {}
    # Loop through row and make new dictionary
    for row_num in range(2, 9):
        VC_id = sheet.cell(row = row_num, column = 1).value
        table[VC_id] = {}
        for col_num in range(2,10):
            target_id = sheet.cell(row = 1, column = col_num).value
            table[VC_id][target_id] = sheet.cell(row = row_num, column = col_num).value
    return table

# Add vc model id only
def AddingVCModelID(filepath):
    if filepath[-3:].lower() != "wav":
        print(f"{filepath} is not a wav file")
        return
    
    dir, name = os.path.split(filepath)

    newname = name[:11] + f"{VC_id:02}" + name[11:]
    # print(filepath, "->", dir+"/"+newname)
    os.rename(filepath, dir+"/"+newname)
# list_folders_and_files(folderName, AddingVCModelID)

# ExcelSheetToTable("C:/Users/jazzt/Desktop/CCA/Deepfake_Audio_Detection/Mangio-RVC-v23.7.0_INFER_TRAIN/Mangio-RVC/opt/voice_adjustment.xlsx")
# SetSpoofedDataName("C:/Users/jazzt/Desktop/CCA/Deepfake_Audio_Detection/Mangio-RVC-v23.7.0_INFER_TRAIN/Mangio-RVC/opt/New_folder")

sys.setrecursionlimit(100000)
def CopyAllFiles(startingDir, targetDir): # Loop through all folder and obtain the files in all of them, then copy to a designated folder
    for item in os.listdir(startingDir):
        if item == os.path.split(targetDir)[1]: # avoid same folder
            continue
        if os.path.isdir(startingDir+'/'+item):
            print(f"entering {startingDir+'/'+item}")
            CopyAllFiles(startingDir+'/'+item, targetDir)
        else:
            shutil.copy(startingDir+'/'+item, targetDir+'/'+item)

# CopyAllFiles("C:/Users/jazzt/Desktop/CCA/Deepfake_Audio_Detection/Mangio-RVC-v23.7.0_INFER_TRAIN/Mangio-RVC/opt", 
#              "C:/Users/jazzt/Desktop/CCA/Deepfake_Audio_Detection/Mangio-RVC-v23.7.0_INFER_TRAIN/Mangio-RVC/opt/Mangio_RVC_spoofed_data")


# Changing LA naming convention to SG
if 0:
    metapath = "./data/LA/LA/ASVspoof2019_LA_cm_protocols/ASVspoof2019.LA.cm.dev.trl.txt"
    oldfolderPath = "./data/LA/LA/ASVspoof2019_LA_dev/flac"
    newfolderPath = "./data/newLA/LAdev/"

    with open(metapath, "r") as file:
        for line in file:
            info = line.split(' ')
            name = info[1]
            if 'bonafide' in info[4]:
                name = name[5:]
            shutil.copy(os.path.join(oldfolderPath, info[1] + ".flac"),
                        os.path.join(newfolderPath, name + ".wav"))

# Changing LA naming convention to SG, for ASVspoof2021 data, all will be used for eval
if 0:
    metapath = "./data/ASV2021/keys/DF/CM/trial_metadata.txt"
    oldfolderPath = "./data/ASV2021/ASVspoof2021_DF_eval/flac"
    newfolderPath = "./data/ASV2021/ASVspoof2021_DF_eval/flac"

    bonafideCount = 0
    spoofCount = 0

    with open(metapath, "r") as file:
        for line in file:
            info = line.split(' ')
            name = info[1]
            if 'bonafide' in info[5]:
                name = name[5:]
                bonafideCount += 1
            else:
                spoofCount += 1
            # os.rename(os.path.join(oldfolderPath, info[1] + ".flac"),
            #           os.path.join(newfolderPath, name + ".wav"))
        print(f"bonafideCount: {bonafideCount}")
        print(f"spoofCount: {spoofCount}")

