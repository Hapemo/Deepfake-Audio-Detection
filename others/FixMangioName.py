import os
import openpyxl
import shutil
import random
import sys
folderName = "C:/Users/jazzt/Desktop/CCA/Deepfake_Audio_Detection/Mangio-RVC-v23.7.0_INFER_TRAIN/Mangio-RVC/audios"

def list_folders_and_files(dir_path, funcPtr): # the function pointer should only take in the file path
    # Loop through all files and directories in the given path
    for item in os.listdir(dir_path):
        # Get the full path of the item
        item_path = dir_path + "/" + item
        
        if os.path.isdir(item_path):
            # If the item is a directory, recursively list its contents
            print(f"Entering folder: {item_path}")
            list_folders_and_files(item_path, funcPtr)
        else:
            # If the item is a file, print its path
            funcPtr(item_path)

def FixAudioFileName(filepath):
    if filepath[-3:].lower() != "wav":
        print(f"{filepath} is not a wav file")
        return
    dir, name = os.path.split(filepath)
    
    # fix first error, removing SPEAKER<number>
    if (not name[0].isdigit()):
        name = name[11:]
        
    
    # fix second error, where there are more than one .wav in the name, or wav is WAV
    wavOccurance = name.lower().count(".wav")
    while (wavOccurance != 0):
        name = name[:-4]
        wavOccurance-=1
    name = name + ".wav"

    # # fix third error, where there are more than 8 digits in the number
    # if len(name) > 12:
    #     name = name[1:]
    
    print(dir+"/"+name)
    os.rename(filepath, dir+"/"+name)

VCModelID = 00
# Mangio-RVC data didn't have 0 in the name, this is to 
def AddingVCModelID(filepath):
    if filepath[-3:].lower() != "wav":
        print(f"{filepath} is not a wav file")
        return
    dir, name = os.path.split(filepath)


# list_folders_and_files(folderName, FixAudioFileName)

VC_id = 1

def SetSpoofedDataName(startingPath):
    pitchTable = ExcelSheetToTable(startingPath + "/voice_adjustment.xlsx")
    print(startingPath + "/voice_adjustment.xlsx")
    for subfolder in os.listdir(startingPath):
        if not os.path.isdir(startingPath+'/'+subfolder):
            print(subfolder)
            continue
        # Get the full path of the item
        VCmodel_id = subfolder[7:11]
        for fromFolder in os.listdir(startingPath+'/'+subfolder):
            target_id = fromFolder[-4:]
            for name in os.listdir(startingPath+'/'+subfolder+'/'+fromFolder):
                newname = VCmodel_id + name
                newname = newname[:8] + pitchTable[VCmodel_id][target_id] + f"{VC_id:02}" + newname[8:]
                print(newname)

                os.rename(startingPath+'/'+subfolder+'/'+fromFolder+'/'+name, 
                          startingPath+'/'+subfolder+'/'+fromFolder+'/'+newname)

def ExcelSheetToTable(excelSheetPath): # converting excelsheet to a 2d array, 
    sheet = openpyxl.load_workbook(excelSheetPath)['Sheet1']
    
    table = {}
    # Loop through row and make new dictionary
    for row_num in range(2, 12):
        VC_id = sheet.cell(row = row_num, column = 1).value
        table[VC_id] = {}
        for col_num in range(2,10):
            target_id = sheet.cell(row = 1, column = col_num).value
            table[VC_id][target_id] = sheet.cell(row = row_num, column = col_num).value
    return table